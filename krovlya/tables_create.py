from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Товары"

# ==========================================
# ОПРЕДЕЛЯЕМ СТОЛБЦЫ
# ==========================================
columns = [
    # (Заголовок, ширина, комментарий, обязательное)
    ('name', 45, 'Название товара', True),
    ('slug', 45, 'URL-адрес (латиница, дефисы)', True),
    ('category_slug', 22, 'Slug категории из списка', True),
    ('manufacturer_slug', 22, 'Slug производителя из списка', False),
    ('short_description', 50, 'Краткое описание (1-2 предложения)', False),
    ('description', 60, 'Полное описание', False),
    ('price', 12, 'Цена (число)', True),
    ('old_price', 12, 'Старая цена (число)', False),
    ('price_unit', 10, 'м² / шт. / лист / п.м.', False),
    ('available', 10, '1 = в наличии, 0 = нет', False),
    ('stock', 10, 'Количество на складе', False),
    ('brand', 18, 'Бренд', False),
    ('material', 30, 'Материал', False),
    ('color', 15, 'Код цвета RAL', False),
    ('color_name', 20, 'Название цвета', False),
    ('coating', 20, 'Покрытие', False),
    ('coating_thickness', 18, 'Толщина покрытия', False),
    ('thickness', 15, 'Толщина металла', False),
    ('country_brand', 15, 'Страна бренда', False),
    ('country_production', 18, 'Страна производства', False),
    ('purpose', 30, 'Назначение', False),
    ('length', 20, 'Длина', False),
    ('width_total', 15, 'Ширина общая', False),
    ('width_working', 18, 'Ширина рабочая', False),
    ('wave_height', 15, 'Высота волны', False),
    ('step_height', 15, 'Высота ступени', False),
    ('step_pitch', 15, 'Шаг ступени', False),
    ('roof_angle', 15, 'Угол кровли', False),
    ('profile_type', 22, 'Тип профиля', False),
    ('form_type', 20, 'Форма/серия', False),
    ('surface_gloss', 15, 'Глянец/мат', False),
    ('surface_texture', 18, 'Текстура поверхности', False),
    ('back_side', 20, 'Обратная сторона', False),
    ('protective_layer', 22, 'Защитный слой', False),
    ('warranty_appearance', 20, 'Гарантия на внешний вид', False),
    ('warranty_technical', 22, 'Техническая гарантия', False),
    ('resistance', 18, 'Устойчивость', False),
]

# ==========================================
# СТИЛИ
# ==========================================
# Обязательные поля — зелёный фон
fill_required = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
# Необязательные — голубой фон
fill_optional = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
# Строка с комментариями — серый
fill_comment = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
# Строка с примерами — жёлтый
fill_example = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

font_header = Font(name='Arial', size=11, bold=True)
font_comment = Font(name='Arial', size=9, italic=True, color='666666')
font_example = Font(name='Arial', size=10, color='333333')
font_normal = Font(name='Arial', size=10)

align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_comment = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_normal = Alignment(vertical='top', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

# ==========================================
# СТРОКА 1 — ЗАГОЛОВКИ
# ==========================================
for col_idx, (header, width, comment, required) in enumerate(columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = font_header
    cell.alignment = align_header
    cell.border = thin_border
    cell.fill = fill_required if required else fill_optional

    # Ширина столбца
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ==========================================
# СТРОКА 2 — КОММЕНТАРИИ / ПОДСКАЗКИ
# ==========================================
for col_idx, (header, width, comment, required) in enumerate(columns, start=1):
    label = '✅ ' + comment if required else comment
    cell = ws.cell(row=2, column=col_idx, value=label)
    cell.font = font_comment
    cell.alignment = align_comment
    cell.fill = fill_comment
    cell.border = thin_border

# ==========================================
# СТРОКА 3 — ПРИМЕР ТОВАРА 1 (Металлочерепица)
# ==========================================
example1 = {
    'name': 'Металлочерепица "Квинта плюс" 0,5 Satin RAL 8017 шоколад',
    'slug': 'metallocherepitsa-kvinta-plus-satin-ral-8017',
    'category_slug': 'metallocherepitsa',
    'manufacturer_slug': 'grand-line',
    'short_description': 'Оригинальный рисунок профиля металлочерепицы Kvinta plus.',
    'description': 'Металлочерепица Kvinta plus изготовлена из оцинкованной стали.',
    'price': 850,
    'old_price': '',
    'price_unit': 'м²',
    'available': 1,
    'stock': 500,
    'brand': 'Grand Line',
    'material': 'оцинкованная сталь',
    'color': 'RAL 8017',
    'color_name': 'шоколад',
    'coating': 'Satin',
    'coating_thickness': '25 мкм',
    'thickness': '0,5 мм',
    'country_brand': 'Россия',
    'country_production': 'Россия',
    'purpose': 'для кровли',
    'length': '0,47 - 8,00 м',
    'width_total': '1210 мм',
    'width_working': '1150 мм',
    'wave_height': '20 мм',
    'step_height': '30 мм',
    'step_pitch': '350 мм',
    'roof_angle': 'от 12°',
    'profile_type': 'Металлочерепица',
    'form_type': 'Kvinta plus',
    'surface_gloss': 'глянцевая',
    'surface_texture': 'гладкая',
    'back_side': 'эпоксидная серая',
    'protective_layer': 'Zn 100-140 г/м²',
    'warranty_appearance': '10 лет',
    'warranty_technical': '20 лет',
    'resistance': 'к УФ/UV3',
}

for col_idx, (header, width, comment, required) in enumerate(columns, start=1):
    value = example1.get(header, '')
    cell = ws.cell(row=3, column=col_idx, value=value)
    cell.font = font_example
    cell.alignment = align_normal
    cell.fill = fill_example
    cell.border = thin_border

# ==========================================
# СТРОКА 4 — ПРИМЕР ТОВАРА 2 (Сайдинг)
# ==========================================
example2 = {
    'name': 'Сайдинг виниловый D4.5 Docke Premium пломбир',
    'slug': 'sayding-vinilovyj-docke-premium-plombir',
    'category_slug': 'sayding',
    'manufacturer_slug': 'docke',
    'short_description': 'Виниловый сайдинг премиум-класса.',
    'description': '',
    'price': 320,
    'old_price': '',
    'price_unit': 'шт.',
    'available': 1,
    'stock': 1000,
    'brand': 'Docke',
    'material': 'ПВХ',
    'color': '',
    'color_name': 'пломбир',
    'coating': '',
    'coating_thickness': '',
    'thickness': '',
    'country_brand': 'Германия',
    'country_production': 'Россия',
    'purpose': 'для фасада',
    'length': '3660 мм',
    'width_total': '240 мм',
    'width_working': '',
    'wave_height': '',
    'step_height': '',
    'step_pitch': '',
    'roof_angle': '',
    'profile_type': 'Сайдинг',
    'form_type': 'D4.5',
    'surface_gloss': '',
    'surface_texture': '',
    'back_side': '',
    'protective_layer': '',
    'warranty_appearance': '25 лет',
    'warranty_technical': '',
    'resistance': '',
}

for col_idx, (header, width, comment, required) in enumerate(columns, start=1):
    value = example2.get(header, '')
    cell = ws.cell(row=4, column=col_idx, value=value)
    cell.font = font_example
    cell.alignment = align_normal
    cell.fill = fill_example
    cell.border = thin_border

# ==========================================
# СТРОКИ 5-54 — ПУСТЫЕ ДЛЯ ЗАПОЛНЕНИЯ
# ==========================================
for row in range(5, 55):
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=row, column=col_idx, value='')
        cell.font = font_normal
        cell.alignment = align_normal
        cell.border = thin_border

# ==========================================
# ЛИСТ 2 — СПРАВОЧНИК КАТЕГОРИЙ
# ==========================================
ws_cat = wb.create_sheet('Категории (справочник)')

cat_headers = ['slug (вставлять в category_slug)', 'Название', 'Родительская']
cat_data = [
    ('metallocherepitsa', 'Металлочерепица', 'Кровля'),
    ('profnastil', 'Профнастил', 'Кровля'),
    ('gibkaya-cherepitsa', 'Гибкая черепица', 'Кровля'),
    ('kompozitnaya-cherepitsa', 'Композитная черепица', 'Кровля'),
    ('faltsevaya-krovlya', 'Фальцевая кровля', 'Кровля'),
    ('sayding', 'Сайдинг', 'Фасад'),
    ('fasadnye-paneli', 'Фасадные панели', 'Фасад'),
    ('planken', 'Планкен', 'Фасад'),
    ('vodostok', 'Водосток', '—'),
    ('ograzhdeniya', 'Ограждения', '—'),
    ('mineralnaya-vata', 'Минеральная вата', 'Утепление'),
    ('xps', 'Экструдированный пенополистирол', 'Утепление'),
    ('pir-plity', 'PIR-плиты', 'Утепление'),
    ('pilomaterialy', 'Пиломатериалы', '—'),
    ('okna', 'Окна', '—'),
    ('stroyhimiya', 'Стройхимия', '—'),
]

for col_idx, h in enumerate(cat_headers, start=1):
    cell = ws_cat.cell(row=1, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_required
    cell.border = thin_border

for row_idx, (slug, name, parent) in enumerate(cat_data, start=2):
    ws_cat.cell(row=row_idx, column=1, value=slug).font = font_normal
    ws_cat.cell(row=row_idx, column=2, value=name).font = font_normal
    ws_cat.cell(row=row_idx, column=3, value=parent).font = font_normal
    for c in range(1, 4):
        ws_cat.cell(row=row_idx, column=c).border = thin_border

ws_cat.column_dimensions['A'].width = 35
ws_cat.column_dimensions['B'].width = 35
ws_cat.column_dimensions['C'].width = 20

# ==========================================
# ЛИСТ 3 — СПРАВОЧНИК ПРОИЗВОДИТЕЛЕЙ
# ==========================================
ws_mfr = wb.create_sheet('Производители (справочник)')

mfr_headers = ['slug (вставлять в manufacturer_slug)', 'Название', 'Страна']
mfr_data = [
    ('grand-line', 'Grand Line', 'Россия'),
    ('tehnonikol', 'Технониколь', 'Россия'),
    ('metal-profile', 'Metal Profile', 'Россия'),
    ('docke', 'Docke', 'Германия'),
    ('aquasystem', 'AquaSystem', 'Россия'),
    ('rockwool', 'Rockwool', 'Дания'),
    ('penoplex', 'Пеноплэкс', 'Россия'),
    ('velux', 'Velux', 'Дания'),
    ('fakro', 'Fakro', 'Польша'),
    ('ceresit', 'Ceresit', 'Германия'),
    ('paroc', 'Paroc', 'Финляндия'),
    ('ruukki', 'Ruukki', 'Финляндия'),
]

for col_idx, h in enumerate(mfr_headers, start=1):
    cell = ws_mfr.cell(row=1, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_required
    cell.border = thin_border

for row_idx, (slug, name, country) in enumerate(mfr_data, start=2):
    ws_mfr.cell(row=row_idx, column=1, value=slug).font = font_normal
    ws_mfr.cell(row=row_idx, column=2, value=name).font = font_normal
    ws_mfr.cell(row=row_idx, column=3, value=country).font = font_normal
    for c in range(1, 4):
        ws_mfr.cell(row=row_idx, column=c).border = thin_border

ws_mfr.column_dimensions['A'].width = 40
ws_mfr.column_dimensions['B'].width = 25
ws_mfr.column_dimensions['C'].width = 15

# ==========================================
# ВЫСОТА СТРОК
# ==========================================
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 40
ws.row_dimensions[3].height = 25
ws.row_dimensions[4].height = 25

# Закрепляем заголовок
ws.freeze_panes = 'A3'

# ==========================================
# СОХРАНЯЕМ
# ==========================================
filename = 'products_template.xlsx'
wb.save(filename)
print(f"Файл '{filename}' создан!")
print(f"  Лист 1: Товары — заголовки + 2 примера + 50 пустых строк")
print(f"  Лист 2: Справочник категорий ({len(cat_data)} шт.)")
print(f"  Лист 3: Справочник производителей ({len(mfr_data)} шт.)")
print(f"")
print(f"Заполните лист 'Товары' и запустите import_from_csv.py")